from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

def xlsx2json(xlsx):
    members = []

    wb = load_workbook(xlsx)
    ws = wb.active

    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))

    for row in range(2, last_row + 1):
        member = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                member[ws[column_letter + str(1)].value] = str(ws[column_letter + str(row)].value).strip()
        members.append(member)
        
    data = json.dumps(members, sort_keys=True, indent=4)
    with open('data.json', 'w', encoding='utf-8') as f:
        f.write(data)

def json2dict():
    with open('data.json') as j:
        dictMembers = json.load(j)
    return dictMembers

if __name__ == "__main__":
    xlsx2json('Κατάσταση-ενεργών-μελών-Δεκέμβριος-2021.xlsx')
    members = json2dict()
    for member in range(len(members)):
        print("Το μέλος ονομάζεται {} {} (email: {}).\n".format(members[member]['Όνομα'],
         members[member]['Επώνυμο'], members[member]['Email']))